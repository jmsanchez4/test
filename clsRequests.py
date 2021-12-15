import csv
import re
import openpyxl
import settings
import json
from collections import namedtuple


#Option = namedtuple('Option', ['letter','label', 'callback'])
class Menu(object):

	SEPARATOR = '-'


	def __init__(self,title, options,lst=None):
		self._title = ''
		self._options = []
		self._title = title

		Option = namedtuple('Option', ['letter', 'label', 'callback'])
		for option in options:
			self._options.append(Option(option[0], option[1],option[2]))

	def header(self, text):
		
		line = self.SEPARATOR * (len(text) + 2)
		return "\n\t{}\n \t{}\n".format(line,text)

	def display(self):
		string = self.header(self._title)
		string += "\t"

		for option in self._options:
			string += "{}, ".format(option.label)

		return string

	def callback(self, i):
		for x in self._options:
			if str(i).lower() == str(x[0]).lower():
				return x.callback
		# if i <= len(self._options):
		# 	return self._options[i - 1].callback


class MainMenu(object):
	#Option = namedtuple('Option', ['label', 'callback'])
	SEPARATOR = '-'

	#_title = ''
	#_options = []

	def __init__(self, title, options):
		self._title = title
		self._options=[]
		Option = namedtuple('Option', ['label', 'callback'])

		for option in options:
			self._options.append(Option(option[0], option[1]))

	def header(self, text):
		line = self.SEPARATOR * (len(text) + 2)
		return "\t\t{}\n \t\t{}\n\t\t{}\n".format(line,text,line)

	def display(self):
		string = self.header(self._title)

		for i, option in enumerate(self._options):
			string += "\t\t{} {}\n".format(i+1,option.label)

		return string

	def callback(self, i):
		if i <= len(self._options):
			return self._options[i - 1].callback


class RequestWorkbook(object):
	def __init__(self, filename=''):


		self.killsheets = ['Instructions', 'INDEX', 'lists']
		self.wb = ''
		self.ws = ''
		self.sheetnames = ''
		self.data = []

		if filename=='':
			self.filename = raw_input("Path and filename to the requests xlsx file:\n>")
		else:
			self.filename = filename

		try:
			wb = openpyxl.load_workbook(self.filename, data_only=True)
			sheetnames = wb.get_sheet_names()
			self.wb = wb
			self.sheetnames = sheetnames
		except IOError:
			print "\n* File does not exist! Please check and try again!\n" % filename

		for sheet in self.sheetnames:
			if sheet in self.killsheets:
				self.sheetnames.remove(sheet)

			#self.wksheet_menu()

	def __str__(self):
		return str(vars(self))

	def wksheet_menu(self):

		c = False
		while True:
			choice = ''
			sheet = ''
			print "\n" * 100
			print "\t***************************************"
			print "\t***  REQUESTS WORKBOOK OPERATIONS   ***"
			print "\t***************************************"
			print "\n"
			print '\tChoose a selection by number: '
			print "\n"

			# Enumerate the worksheets and ask which one to work
			for count, sheet in enumerate(self.sheetnames, start=1):
				print "\t{}) {}".format(count, sheet)
			choice = raw_input("\n\tPick item to work or (m)ain menu: ")

			#try:
			if str(choice).lower() == 'm':
				return
			try:
				if int(choice) in range(1, len(self.sheetnames) + 1):
					# Get worksheet data based on users choice
					for count, sheet in enumerate(self.sheetnames, start=1):
						if str(count) == str(choice):
							ws = self.wb.get_sheet_by_name(sheet)
							self.ws=ws
				else:
					continue
			except:
				continue

			for index, row in enumerate(self.ws.iter_rows()):
				if index == 0 or index == 1:
					continue
				if row[0].value is not None:
					# row
					# configs.append([cell.value for cell in row])
					self.data.append([(u"" if cell.value is None else unicode(cell.value)) for cell in row])
			if len(self.data) == 0:
				print "No data in chosen worksheet."
				continue
		return self.data

	def killsheet_add(self, sheetname):
		self.killsheets.append(sheetname)

	def killsheet_remove(self, sheetname):
		self.killsheets.remove(sheetname)

	def killsheets_get(self):
		return self.killsheets

	def kilsheets_show(self):
		print self.killsheets


class RequestsHandler(object):
	def __init__(self, filename=''):
		self.requests_list = []
		self.configs_list = []
		self.requests_count = len(self.requests_list)
		self.configs_count = 0


	def __str__(self):
		return str(vars(self))


	def import_requests(self,filename=''):
		requests = RequestWorkbook(filename)
		self.data=requests.data
		while True:

			requests.wksheet_menu()
			# Create list from each row of data and add that list to the configs list configs=[[worksheet row1],[worksheet row 2],[worksheet row etc.etc.]]
			if requests.ws.title == 'LTM & GTM Load Balancing':
				self.build_request_objects()
				break
			elif requests.ws.title == 'Migrate VS':
				pass
			elif requests.ws.title == 'LTM Pool Members':
				pass
			else:
				continue


	def build_request_objects(self):
		newreq = {}
		for i in self.data:
			# initialize emtpy payload dictionary. We will fill it and create request objects

			env = str(i[1]).strip()

			newreq['service_owner'] = i[0].strip()
			newreq['env'] = str(i[1]).strip()
			newreq['vs_name'] = i[2].strip()
			newreq['vs_description'] = i[3].strip() + " Service Owner: " + newreq['service_owner']
			newreq['vs_ip'] = i[4].strip()
			newreq['vs_port'] = str(i[5]).strip()
			newreq['vs_protocol'] = i[6].strip().lower()
			newreq['vs_redirection'] = i[7].strip()
			newreq['vs_persistance_profile'] = i[8]
			newreq['pool_name'] = i[9].strip()
			newreq['pool_description'] = i[10].strip()
			newreq['pool_lb_method'] = i[11].strip()

			if i[12] == '':
				newreq['pool_monitor'] = ''
			else:
				newreq['pool_monitor'] = i[12]

			newreq['ssl'] = i[13].strip()
			newreq['wideip'] = i[14].strip().lower()
			newreq['wideip_lb_method'] = i[15].strip()
			newreq['gtm_pool_lb_method'] = i[16].strip()

			if not i[17]:
				i[17] = 0
			newreq['gtm_order'] = i[17]

			if env == 'ICP_CAMPUS':
				try:
					newreq['pna_prod_vip'] = i[19].strip()
					newreq['crn_prod_vip'] = i[21].strip()
					newreq['pna_prod_campus_snat'] = i[27].strip()
					newreq['crn_prod_campus_snat'] = i[28].strip()

				except:
					print "\n!!! Required values were not provided for the Campus IP Adresses !!!\n"
					exit()

			if env == 'EICP_APP':
				try:
					newreq['pna_prod_vip'] = i[19].strip()
					newreq['pna_nprod_vip'] = i[20].strip()
					newreq['crn_prod_vip'] = i[21].strip()
					newreq['crn_nprod_vip'] = i[22].strip()

				except:
					print "\n!!! Required values were not provided for the APP Tier  IP Adresses !!!\n"
					exit()

			if env == "EICP_WEB":
				try:
					newreq['pna_prod_vip'] = i[19].strip()
					newreq['pna_nprod_vip'] = i[20].strip()
					newreq['crn_prod_vip'] = i[21].strip()
					newreq['crn_nprod_vip'] = i[22].strip()
					newreq['pna_waf_vip'] = i[23].strip()
					newreq['pna_waf_np_vip'] = i[24].strip()
					newreq['crn_waf_vip'] = i[25].strip()
					newreq['crn_waf_np_vip'] = i[26].strip()
				except:
					print "\n!!! Required values were not provided for the WEB Tier  IP Adresses !!!\n"
					exit()

			req = F5Request(newreq)
			self.requests_list.append(req)
		self.requests_count=len(self.requests_list)


	def enumerate_requests(self):
		return [n for n in enumerate(self.requests_list, start=1)]


	def get_requests(self):
		return self.requests_list


	def get_request(self,idx):
		for c, request in enumerate(self.requests_list, start=1):
			if str(c) == str(idx):
				return request


	def generate_all_request_configs(self):
		for request in self.requests_list:
			for config in request.configs:
				if config not in self.configs_list:
					self.configs_list.append(config)


	def get_configs_list(self):
		return self.configs_list


	def config_all(self):
		pass

	def requests_count(self):
		rc= len(self.requests_list)
		return rc

class F5Connect(object):

	"""
	F5 connection object and device specific utilities
	"""

	def __init__(self, user=None, passwd=None, host=None, partition="Common"):
		self.user = user
		self.password = passwd
		self.host = host
		self.bigip = self.build_management_root()
		self.b = self.build_request_object()
		self.partition = partition
		self.hostname = ''
		self.get_hostname()
		self.cli_user = ''
		self.cli_password = ''

	def build_management_root(self):

		"""
		return a connection object using F5 SDK library
		:return:
		"""

		try:

			bigip = ManagementRoot(self.host,self.user,self.password)

			return bigip

		except Exception, e:
			# print Exception.args
			print("\t%s" % e)
			print(
				"Error: -- \n Please make sure environment variables for BigIP IP, Username and Password are set correctly!")
			bigip = False

			return

	def build_request_object(self):

		"""
		returns a connection object using the requests and json library

		:return:
		"""

		b = rq.session()
		b.auth = (self.user, self.password)
		b.verify = False
		b.headers.update({'Content-Type': 'application/json'})
		b.b_base = "https://%s/mgmt/tm" % self.host
		return b

	def build_sftp_object(self):

		"""
		Uses paramiko tp scp or sftp the CSR down to the csr_requests folder
		:param file:
		"""

		#Get the hostname this will be downloaded from
		hostname = self.hostname

		# Open a transport
		host, port = self.host, 22
		transport = paramiko.Transport((host, port))

		# Auth
		username, password = self.user, self.password
		transport.connect(None, username, password)

		# Go!
		sftp = paramiko.SFTPClient.from_transport(transport)

		return sftp, transport

	def build_ssh_object(self):
		"""
		:return:
		"""



		# Open a transport
		host, port = self.host, 22

		# Auth
		username, password = self.cli_user, self.cli_password

		ssh = paramiko.SSHClient()
		ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
		ssh.connect(host,port,username,password)

		return ssh

	def get_hostname(self):

		"""
		Gets the hostname of self
		:return:
		"""


		settings = self.bigip.tm.sys.global_settings.load()
		hostname = settings.hostname.split('.')[0]
		self.hostname = hostname

		return

	def get_active(self):

		"""
		This attempts to show state and synch state. Needs further development so
		script can determine which device is active and the synchronization status
		so it can eventually only configure the active device and synch form it
		"""
		#TODO - Function to determine synch, active device, and synch from active device

		state = self.bigip.tm.cm.devices.get_collection()
		synch = self.bigip.tm.cm.sync_status

		print state[0].failoverState
		print synch.name

	def __get_creds__(self):

		"""
		private methoed prompts for username and password and returns same
		:return:
		"""

		username = raw_input("Username: ")
		password = getpass.getpass(stream=sys.stderr)

		return username, password

class F5LTM(F5Connect):
	def __init__(self, d_ip='', username='', password=''):

		# Initialize default attributes
		self.payload = {}
		self.d_ip = d_ip
		self.d_name = ''
		self.d_location = ''







	def __str__(self):
		return str(vars(self))

	def get_pool(self):
		pool = self.bigip.tm.ltm.pools.pool.load(name=self.pool_name, partition=self.partition)
		return pool

	def create_pool(self):

		if self.bigip.tm.ltm.pools.pool.exists(name=self.pool_name):
			print ("LTM Pool >> {} << exist ... skipping !!! ".format(self.pool_name))
			pass
		else:
			self.bigip.tm.ltm.pools.pool.create(**self.pool_payload)
			return

	def create_fqdn_pool(self):
		# Create a FQDN node first if doesn't exists
		if self.bigip.tm.ltm.nodes.node.exists(name=self.pool_members[0]['tmName']):
			print ("Node >> {} << exist ... skipping NODE creation  !!! ".format(self.pool_members[0]['tmName']))
		else:
			self.bigip.tm.ltm.nodes.node.create(name=self.pool_members[0]['tmName'], address='any6',
													fqdn={'tmName': self.pool_members[0]['tmName']},
													partition=self.partition)

		# Create a pool if it doesn't exists
		if self.bigip.tm.ltm.pools.pool.exists(name=self.pool_name):
			print ("Pool >> {} << exist ... skipping !!! ".format(self.pool_name))
		else:
			self.bigip.tm.ltm.pools.pool.create(**self.pool_payload)
			print ("Pool >> {} << Created !!! ".format(self.pool_name))

	def del_pool(self):
		if self.bigip.tm.ltm.pools.pool.exists(name=self.pool_name, partition=self.partition):
			pool_b = self.bigip.tm.ltm.pools.pool.load(name=self.pool_name, partition=self.partition)
			try:
				x = pool_b.delete()
			except Exception, e:
				print("\t%s" % e)
		else:
			print "pool does not exist..."

	def get_pools(self):
		pools = self.bigip.tm.ltm.pools.get_collection()
		for pool in pools:
			for member in pool.members_s.get_collection():
				print member.name

	def add_pool_member(self):
		x = self.b.post('%s/ltm/pool/~Common~%s/members' % (self.b.b_base, self.b.poolname),
						data=json.dumps(self.payload)).json()
		print(json.dumps(x, indent=4, sort_keys=True))

	def del_node(self):
		if self.bigip.tm.ltm.nodes.node.exists(name=self.nodename, partition=self.partition):
			node = self.bigip.tm.ltm.nodes.node.load(name=self.nodename, partition=self.partition)
			x = node.delete()

	def create_snatpool(self):

		try:
			self.bigip.tm.ltm.snatpools.snatpool.create(**self.vs_snat_payload)

		except Exception, e:
			print("\t%s" % e)

	def del_snatpool(self):

		try:
			self.snatpool = self.bigip.tm.ltm.snatpools.snatpool.load(partition=self.partition,
																		  name=self.vs_snatpoolname)
			print self.snatpool.raw
			self.snatpool.delete()
			print "Snat Pool Deleted- %s" % self.vs_snatpoolname
		except Exception, e:
				print("\t%s" % e)

	def create_virtual(self):
		if self.bigip.tm.ltm.virtuals.virtual.exists(partition=self.partition, name=self.vs_name):
			print ("Virtual >> {} << exist ... skipping !!! \n".format(self.vs_name))
		# logging.info("Virtual >> {} << exist ... skipping !!! \n".format(virt_payload['name']))
		else:
			self.bigip.tm.ltm.virtuals.virtual.create(**self.vs_payload)
			print ("Virtual >> {} << Created !!!".format(self.vs_payload))

	def delete_virtual(self):
		print ("\nDeleting Virtual {}".format(self.vs_name))
		if self.bigip.tm.ltm.virtuals.virtual.exists(partition=self.partition, name=self.vs_name):
			virtual = self.bigip.tm.ltm.virtuals.virtual.load(partition='Common', name=self.vs_name)
			virtual.delete()
			print ("Virtual >> {} << Deleted !!! ".format(self.vs_name))
		else:
			print("Virtual server does not exist")

		# Move to a collection class
		# def get_vs_all(self):
		#
		#
		# # CAPTURE LIST OF CLIENT SSL PROFILES
		# client_ssls = self.bigip.tm.ltm.profile.client_ssls.get_collection()
		# # https:// <F5_mgmt_IP>/mgmt/tm/ltm/profile/client-ssl
		# listClientSsl = []
		# for client_ssl in client_ssls:
		# 	listClientSsl.append(client_ssl.name)
		#
		# # CAPTURE LIST OF SERVER SSL PROFILES
		# server_ssls = self.bigip.tm.ltm.profile.server_ssls.get_collection()
		# # https:// <F5_mgmt_IP>/mgmt/tm/ltm/profile/server-ssl
		# listServerSsl = []
		# for server_ssl in server_ssls:
		# 	listServerSsl.append(server_ssl.name)
		#
		# # CAPTURE VIRTUAL SERVER INFORMATION
		# virtuals = self.bigip.tm.ltm.virtuals.get_collection()
		# # https:// <F5_mgmt_IP>/mgmt/tm/ltm/virtual
		# for count, virtual in enumerate(virtuals, start=1):
		# 	print("------------")
		# 	virtual.count = count
		# 	print "#%s  " % count
		# 	print("Virtual: {}".format(virtual.name))
		# 	print("Destination: {}".format(re.search('[^\/]+$', virtual.destination).group(0)))
		# 	if hasattr(virtual, 'description'):
		# 		print("Description: {}".format(virtual.description))
		# 	else:
		# 		print("Description: None")
		# 	print("Partition: {}".format(virtual.partition))
		# 	if hasattr(virtual, 'subPath'):
		# 		print("SubPath: {}".format(virtual.subPath))
		# 	else:
		# 		print("SubPath: None")
		#
		# 	listClientSsl_inUse = []
		# 	listServerSsl_inUse = []
		# 	for profile in virtual.profiles_s.get_collection():
		# 		# https:// <F5_mgmt_IP>/mgmt/tm/ltm/virtual/<virtual_name>/profiles
		# 		if profile.name in listClientSsl:
		# 			listClientSsl_inUse.append(profile.name)
		# 		if profile.name in listServerSsl:
		# 			listServerSsl_inUse.append(profile.name)
		# 	if listClientSsl_inUse:
		# 		for prof in listClientSsl_inUse:
		# 			print("Client SSL: {}".format(prof))
		# 	else:
		# 		print("Client SSL: None")
		# 	if listServerSsl_inUse:
		# 		for prof in listServerSsl_inUse:
		# 			print("Server SSL: {}".format(prof))
		# 	else:
		# 		print("Server SSL: None")
		# 	if hasattr(virtual, 'rules'):
		# 		for rule in virtual.rules:
		# 			print("Rule: {}".format(re.search('[^\/]+$', rule).group(0)))
		# 	else:
		# 		print("Rule: None")
		# 	if hasattr(virtual, 'persist'):
		# 		for persist in virtual.persist:
		# 			print("Persistence: {}".format(persist['name']))
		# 	else:
		# 		print("Persistence: None")
		# 	if hasattr(virtual, 'pool'):
		# 		print("Pool: {}".format(re.search('[^\/]+$', virtual.pool).group(0)))
		# 		if hasattr(virtual, 'subPath'):
		# 			poolName = virtual.pool.split("/")[3]
		# 			poolSubpath = virtual.pool.split("/")[2]
		# 			poolPartition = virtual.pool.split("/")[1]
		# 			pool = bigip.tm.ltm.pools.pool.load(name=poolName, subPath=poolSubpath, partition=poolPartition)
		# 		# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>
		# 		else:
		# 			poolName = virtual.pool.split("/")[2]
		# 			poolPartition = virtual.pool.split("/")[1]
		# 			pool = bigip.tm.ltm.pools.pool.load(name=poolName, partition=poolPartition)
		# 		# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>
		# 		poolMembers = pool.members_s.get_collection()
		# 		# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>/members
		# 		if poolMembers:
		# 			for member in poolMembers:
		# 				print("Member: {}".format(member.name))
		# 		else:
		# 			print("Member: None")
		# 	else:
		# 		print("Pool: None")
		# 		print("Member: None")
		# print("------------")
		#
		# return virtuals

	def get_vs_object(self):



		try:
			virtual = self.bigip.tm.ltm.virtuals.virtual.load(partition=self.partition, name=self.vs_name,
															 requests_params={'params': 'expandSubcollections=true'})
			return virtual
		except Exception, e:
			print("\t%s" % e)
			return e

	def show_virtual(self):
		try:
			virtual = self.bigip.tm.ltm.virtuals.virtual.load(partition=self.partition, name=self.vs_name,
																  requests_params={
																	  'params': 'expandSubcollections=true'})
		except Exception, e:
			print("\t%s" % e)
			return

		# CAPTURE LIST OF CLIENT SSL PROFILES
		client_ssls = self.bigip.tm.ltm.profile.client_ssls.get_collection()
		# https:// <F5_mgmt_IP>/mgmt/tm/ltm/profile/client-ssl
		listClientSsl = []
		for client_ssl in client_ssls:
			listClientSsl.append(client_ssl.name)

		# CAPTURE LIST OF SERVER SSL PROFILES
		server_ssls = self.bigip.tm.ltm.profile.server_ssls.get_collection()
		# https:// <F5_mgmt_IP>/mgmt/tm/ltm/profile/server-ssl
		listServerSsl = []
		for server_ssl in server_ssls:
			listServerSsl.append(server_ssl.name)

		print("------------")
		print("Virtual: {}".format(virtual.name))
		print("Destination: {}".format(re.search('[^\/]+$', virtual.destination).group(0)))
		print("Partition: {}".format(virtual.partition))
		if hasattr(virtual, 'description'):
			print("Description: {}".format(virtual.description))
		else:
			print("Description: None")
		if hasattr(virtual, 'subPath'):
			print("SubPath: {}".format(virtual.subPath))
		else:
			print("SubPath: None")

		listClientSsl_inUse = []
		listServerSsl_inUse = []
		for profile in virtual.profiles_s.get_collection():
			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/virtual/<virtual_name>/profiles
			if profile.name in listClientSsl:
				listClientSsl_inUse.append(profile.name)
			if profile.name in listServerSsl:
				listServerSsl_inUse.append(profile.name)
		if listClientSsl_inUse:
			for prof in listClientSsl_inUse:
				print("Client SSL: {}".format(prof))
		else:
			print("Client SSL: None")
		if listServerSsl_inUse:
			for prof in listServerSsl_inUse:
				print("Server SSL: {}".format(prof))
		else:
			print("Server SSL: None")
		if hasattr(virtual, 'rules'):
			for rule in virtual.rules:
				print("Rule: {}".format(re.search('[^\/]+$', rule).group(0)))
		else:
			print("Rule: None")
		if hasattr(virtual, 'persist'):
			for persist in virtual.persist:
				print("Persistence: {}".format(persist['name']))
		else:
			print("Persistence: None")

		for profile in virtual.profiles_s.get_collection():
			print("Profile: {}".format(profile.name))

		if hasattr(virtual, 'pool'):
			print("Pool: {}".format(re.search('[^\/]+$', virtual.pool).group(0)))
			if hasattr(virtual, 'subPath'):
				poolName = virtual.pool.split("/")[3]
				poolSubpath = virtual.pool.split("/")[2]
				poolPartition = virtual.pool.split("/")[1]
				pool = self.bigip.tm.ltm.pools.pool.load(name=poolName, subPath=poolSubpath, partition=poolPartition)
			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>
			else:
				poolName = virtual.pool.split("/")[2]
				poolPartition = virtual.pool.split("/")[1]
				pool = self.bigip.tm.ltm.pools.pool.load(name=poolName, partition=poolPartition)

			if hasattr(pool, 'monitor'):
				print("Monitors: {}".format(pool.monitor))
			else:
				print("Monitors: None")

			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>
			poolMembers = pool.members_s.get_collection()
			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>/members
			if poolMembers:
				for member in poolMembers:
					print("Member: {}".format(member.name))
			else:
				print("Member: None")
		else:
			print("Pool: None")
			print("Member: None")
		print("------------")
		print("\n")

	def gtm_reports(self):
		report = []

		try:

			virtual = self.bigip.tm.gtm.wideips.wideip.load(partition=self.partition, name=self.wideip,
														   requests_params={'params': 'expandSubcollections=true'})
		except Exception, e:
			print("\t%s" % e)

		report.append("------------")
		report.append("Source: {}".format(self.bigip.hostname))
		report.append("Virtual: {}".format(virtual.name))
		report.append("Partition: {}".format(virtual.partition))
		if hasattr(virtual, 'description'):
			report.append("Description: {}".format(virtual.description))
		else:
			report.append("Description: None")

		if hasattr(virtual, 'pools'):
			for x in virtual.pools:
				report.append("Pool: {}".format(x['name']))
				poolname = x['name']

			pool = self.bigip.tm.gtm.pools.pool.load(name=poolname, partition='Common')
			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>


			if hasattr(virtual, 'poolLbMode'):
				report.append("poolLbMode: {}".format(virtual.poolLbMode))
			else:
				report.append("poolLbMode: None")

			if hasattr(pool, 'monitor'):
				report.append("Monitors: {}".format(pool.monitor))
			else:
				report.append("Monitors: None")

			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>
			poolMembers = pool.members_s.get_collection()
			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>/members
			if poolMembers:
				for member in poolMembers:
					report.append("Member: {}".format(member.name))
			else:
				report.append("Member: None")
		else:
			report.append("Pool: None")
			report.append("Member: None")
			report.append("------------")

		return report

	def vs_report(self):
		report = []
		try:

			virtual = self.bigip.tm.ltm.virtuals.virtual.load(partition=self.partition, name=self.vs_name,
																  requests_params={
																	  'params': 'expandSubcollections=true'})
		except Exception, e:
			print("\t%s" % e)
			pass

		# CAPTURE LIST OF CLIENT SSL PROFILES
		client_ssls = self.bigip.tm.ltm.profile.client_ssls.get_collection()
		# https:// <F5_mgmt_IP>/mgmt/tm/ltm/profile/client-ssl
		listClientSsl = []
		for client_ssl in client_ssls:
			listClientSsl.append(client_ssl.name)

		# CAPTURE LIST OF SERVER SSL PROFILES
		server_ssls = self.bigip.tm.ltm.profile.server_ssls.get_collection()
		# https:// <F5_mgmt_IP>/mgmt/tm/ltm/profile/server-ssl
		listServerSsl = []
		for server_ssl in server_ssls:
			listServerSsl.append(server_ssl.name)

		report.append("------------")
		report.append("Source: {}".format(self.bigip.hostname))
		report.append("Virtual: {}".format(virtual.name))
		report.append("Destination: {}".format(re.search('[^\/]+$', virtual.destination).group(0)))
		report.append("Partition: {}".format(virtual.partition))
		if hasattr(virtual, 'description'):
			report.append("Description: {}".format(virtual.description))
		else:
			report.append("Description: None")
		if hasattr(virtual, 'subPath'):
			report.append("SubPath: {}".format(virtual.subPath))
		else:
			report.append("SubPath: None")

		listClientSsl_inUse = []
		listServerSsl_inUse = []
		for profile in virtual.profiles_s.get_collection():
			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/virtual/<virtual_name>/profiles
			if profile.name in listClientSsl:
				listClientSsl_inUse.append(profile.name)
			if profile.name in listServerSsl:
				listServerSsl_inUse.append(profile.name)
		if listClientSsl_inUse:
			for prof in listClientSsl_inUse:
				report.append("Client SSL: {}".format(prof))
		else:
			report.append("Client SSL: None")
		if listServerSsl_inUse:
			for prof in listServerSsl_inUse:
				report.append("Server SSL: {}".format(prof))
		else:
			report.append("Server SSL: None")
		if hasattr(virtual, 'rules'):
			for rule in virtual.rules:
				report.append("Rule: {}".format(re.search('[^\/]+$', rule).group(0)))
		else:
			report.append("Rule: None")
		if hasattr(virtual, 'persist'):
			for persist in virtual.persist:
				report.append("Persistence: {}".format(persist['name']))
		else:
			report.append("Persistence: None")

		for profile in virtual.profiles_s.get_collection():
			report.append("Profile: {}".format(profile.name))

		if hasattr(virtual, 'pool'):
			report.append("Pool: {}".format(re.search('[^\/]+$', virtual.pool).group(0)))
			if hasattr(virtual, 'subPath'):
				poolName = virtual.pool.split("/")[3]
				poolSubpath = virtual.pool.split("/")[2]
				poolPartition = virtual.pool.split("/")[1]
				pool = self.bigip.tm.ltm.pools.pool.load(name=poolName, subPath=poolSubpath,
															 partition=poolPartition)
			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>
			else:
				poolName = virtual.pool.split("/")[2]
				poolPartition = virtual.pool.split("/")[1]
				pool = self.bigip.tm.ltm.pools.pool.load(name=poolName, partition=poolPartition)

			if hasattr(pool, 'monitor'):
				report.append("Monitors: {}".format(pool.monitor))
			else:
				report.append("Monitors: None")

			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>
			poolMembers = pool.members_s.get_collection()
			# https:// <F5_mgmt_IP>/mgmt/tm/ltm/pool/<pool_name>/members
			if poolMembers:
				for member in poolMembers:
					report.append("Member: {}".format(member.name))
			else:
				report.append("Member: None")
		else:
			report.append("Pool: None")
			report.append("Member: None")
			report.append("------------")

		return report

	def menu(self):
		print self
		print "Options:" \
			  "1) Create Pool" \
			  "2) Create Virtual" \
			  "3) Delete Pool" \
			  "4) Delete Virtual" \
			  "5) Report "

		choice=raw_input("Choice?")

class F5LTMConfig(F5Controller):
	def __init__(self):
		# Initialize payload dictionaries
		self.d_ip = d_ip
		self.d_name = ''
		self.d_type = ''
		self.d_location = ''
		self.d_prod = ''
		self.d_device_group=''

		self.vs_payload = {}
		self.vs_name = ''
		self.partition = 'Common'
		self.vs_description = ''
		self.vs_ip = ''
		self.vs_port = ''
		self.vs_destination = self.vs_ip + ':' + self.vs_port
		self.vs_redirect = ''
		self.vs_persistance = ''

		self.vs_snat_payload = {}
		self.vs_snatpool = ''
		self.vs_snatpoolname = ''

		self.pool_name = ''
		self.pool_description = ''
		self.pool_loadBalancingMode = ''
		self.pool_monitor = ''
		self.pool_members = []

		self.pool_payload = {'name': self.pool_name, 'description': self.pool_description,
							 'loadBalancingMode': self.pool_loadBalancingMode, 'monitor': self.pool_monitor, }

		self.nodename = ''
		self.node_payload = {}




	def __str__(self):
		return str(vars(self))

	def create_pool(self):
		# self.pool_payload=vars(self)
		self.pool_payload = {'name': self.pool_name, 'description': self.pool_description,
							 'loadBalancingMode': self.pool_loadBalancingMode, 'monitor': self.pool_monitor, }

		# create LTM pools payload

		# controllerLTM.Create_Pool(self.bigip, self.pool_payload)
		super(F5LTMConfig, self).create_pool()

	def del_pool(self):
		self.pool_payload = {'name': self.pool_name}

		super(F5LTMConfig, self).del_pool()

	def create_virtual(self):
		self.vs_payload = {'name': self.vs_name, 'description': self.vs_description, 'destination': self.vs_destination,
						   'ipProtocol': self.vs_protocol, 'pool': self.pool_name,

						   }
		super(F5LTMConfig, self).create_virtual()


	def obj_menu(self):
		print "{} {} {} {} "


class F5WAFConfig(F5Controller):
	def __init__(self):
		super(F5WAFConfig, self).__init__()


class F5GTMConfig(F5Controller):
	def __init__(self):
		super(F5GTMConfig, self).__init__()


class F5Request(F5Controller):
	# Creates a new request object for pool, ltm, waf, and gtm according to environment.
	# Subclasses LTMConfig and inits from parent to get a set of default values which are updated subsequently by data received from the request.
	# This ensure the controller get the data it needs to execute new creates even if it may not be part of the request.

	def __init__(self, request):
		super(F5Request, self).__init__()

		# set lits to store this requests device configurations
		self.configs = []
		self.configs_ltm = []
		self.configs_waf = []
		self.configs_gtm = []

		# Default values. The controller expects certain items. If we do not find them in the request, the default will be set
		# Request info
		if 'service_owner' in request:
			self.service_owner = request['service_owner']
		if 'env' in request:
			self.env = request['env']
			self.environment = settings.initialize(str(self.env))
		if 'ssl' in request:
			self.ssl = request['ssl']

		# Device info
		if 'd_ip' in request:
			self.d_ip = request['d_ip']
		if 'd_name' in request:
			self.d_name = request['d_name']
		if 'd_location' in request:
			self.d_location = request['d_location'].lower()
		if 'd_prod' in request:
			self.d_prod = request['d_prod'].lower()
		if 'd_type' in request:
			self.d_type = request['d_type'].lower()

		# vs info from request
		if 'vs_name' in request:
			self.vs_name = request['vs_name']
		if 'vs_description' in request:
			self.vs_description = request['vs_description']
		if 'vs_ip' in request:
			self.vs_ip = request['vs_ip']
		if 'vs_port' in request:
			self.vs_port = request['vs_port']
		if 'vs_destination' in request:
			self.vs_destination = request['vs_destination']
		if 'vs_redirect' in request:
			self.vs_redirect = request['vs_redirect']
		if 'vs_persistance' in request:
			self.vs_persistance = request['vs_persistance']

		# pool info from request
		if 'pool_name' in request:
			self.pool_name = request['pool_name']
		if 'pool_description' in request:
			self.pool_description = request['pool_description']
		if 'pool_loadBalancingMode' in request:
			self.pool_loadBalancingMode = request['pool_loadBalancingMode']

		if 'pool_monitor' in request:
			self.pool_monitor = request['pool_monitor']

		# GTM info from request
		if 'gtm_wideip' in request:
			self.gtm_wideip = request['gtm_wideip']
		if 'gtm_order' in request:
			self.gtm_order = request['gtm_order']
		if 'gtm_pool_loadBalancingMode' in request:
			self.gtm_pool_loadBalancingMode = request['gtm_pool_loadBalancingMode']
		if 'gtm_wideip_loadBalancingMode' in request:
			self.gtm_wideip_loadBalancingMode = request['gtm_wideip_loadBalancingMode']

		# ip information for the requested environment from spreadsheet
		if 'pna_prod_vip' in request:
			self.pna_prod_vip = request['pna_prod_vip']
		if 'pna_nprod_vip' in request:
			self.pna_nprod_vip = request['pna_nprod_vip']
		if 'crn_prod_vip' in request:
			self.crn_prod_vip = request['crn_prod_vip']
		if 'crn_nprod_vip' in request:
			self.crn_nprod_vip = request['crn_nprod_vip']
		if 'pna_waf_vip' in request:
			self.pna_waf_vip = request['pna_waf_vip']
		if 'pna_waf_np_vip' in request:
			self.pna_waf_np_vip = request['pna_waf_np_vip']
		if 'crn_waf_vip' in request:
			self.crn_waf_vip = request['crn_waf_vip']
		if 'crn_waf_np_vip' in request:
			self.crn_waf_np_vip = request['crn_waf_np_vip']
		if 'pna_prod_campus_snat' in request:
			self.pna_prod_campus_snat = request['pna_prod_campus_snat']
		if 'crn_prod_campus_snat' in request:
			self.crn_prod_campus_snat = request['crn_prod_campus_snat']
		self.gen_configs()

	def __str__(self):
		return str(vars(self))

	# def _get_environment(self):
	# 	environment = settings.initialize('EICP_WEB')
	# 	print environment

	def _get_ip(self, df5):
		# This gets ip for the configuration object according to location and production status
		# LTM VIP & WAF VIP are assigned according to location and whether they are prod or non-prod. If there is a WAF in the environment, the LTM becomes a pool member on the WAF
		# if it is an ltm, assign it the vip. If it is a waf, assign it the waf vip and assign the ltm vip as the waf pool member ip

		if df5.d_location.lower() == "piscataway" and df5.d_prod.lower() == 'prod':

			if df5.d_type == 'ltm':
				vs_ip = self.pna_prod_vip
			if df5.d_type == 'waf':
				vs_ip = self.pna_waf_vip
				waf_pool_member_ip = self.pna_prod_vip


		elif df5.d_location.lower() == "piscataway" and df5.d_prod.lower() == 'non-prod':

			if df5.d_type == 'ltm':
				vs_ip = self.pna_nprod_vip
			if df5.d_type == 'waf':
				vs_ip = self.pna_waf_np_vip
				waf_pool_member_ip = self.pna_nprod_vip


		elif df5.d_location.lower() == "cranford" and df5.d_prod.lower() == 'prod':

			if df5.d_type == 'ltm':
				vs_ip = self.crn_prod_vip
			if df5.d_type == 'waf':
				vs_ip = self.crn_waf_vip
				waf_pool_member_ip = self.crn_prod_vip


		elif df5.d_location.lower() == "cranford" and df5.d_prod.lower() == 'non-prod':

			if df5.d_type == 'ltm':
				vs_ip = self.crn_nprod_vip

			if df5.d_type == 'waf':
				vs_ip = self.crn_waf_np_vip
				waf_pool_member_ip = self.crn_nprod_vip
		else:

			exit(
				"Could not allocate IP to Device. Please check that the form has the correct IP information in the fields.")

		return vs_ip

	def _gen_ltm_configs(self):
		# Generate ltm configuration objects for this request
		if 'LTM' in self.environment.keys():
			for ltm in self.environment['LTM']:
				ltmconfig = F5LTMConfig()

				# Add device information to config object
				ltmconfig.d_type = 'ltm'
				ltmconfig.d_ip = ltm['ip']
				ltmconfig.d_name = ltm['name'].lower()
				ltmconfig.d_location = str(ltm['location']).lower()
				ltmconfig.d_prod = ltm['prod'].lower()
				ltmconfig.d_gtmpool = ltm['gtmpool'].lower()

				# add virtual server info to config object
				ltmconfig.vs_name = self.vs_name
				ltmconfig.vs_ip = self._get_ip(ltmconfig)  # gets ip from below according to location and whether prod/non-prod
				ltmconfig.destination = ltmconfig.vs_ip + ":" + self.vs_port
				ltmconfig.vs_description = self.vs_description
				ltmconfig.vs_port = self.vs_port
				ltmconfig.vs_redirect = self.vs_redirect
				ltmconfig.vs_persistance = self.vs_persistance

				# add virtual server pool info to config object
				ltmconfig.pool_name = self.pool_name
				ltmconfig.pool_description = self.pool_description
				ltmconfig.pool_loadBalancingMode = self.pool_loadBalancingMode
				ltmconfig.pool_monitor = self.pool_monitor
				# ltmconfig.ltm=ltm #TODO: ltm is a dictionary from environments in settings. Need to class devices and environments

				# uodte self list
				self.configs.append(ltmconfig)
				self.configs_ltm.append(ltmconfig)



		else:
			print "No LTM in this environment"
			return

	def _gen_waf_configs(self):
		if 'WAF' in self.environment.keys():
			for waf in self.environment['WAF']:
				wafconfig = F5WAFConfig()

				# Add device information to config object
				wafconfig.d_type = 'waf'
				wafconfig.d_ip = waf['ip']
				wafconfig.d_name = waf['name'].lower()
				wafconfig.d_location = waf['location'].lower()
				wafconfig.d_prod = waf['prod'].lower()  # TODO: Create non prod name if this is a non-prod device
				wafconfig.d_gtmpool = waf['gtmpool']

				# add virtual server info to config object
				wafconfig.vs_name = self.vs_name
				wafconfig.vs_ip = self._get_ip(wafconfig)  # gets ip according to location and whether prod/non-prod
				wafconfig.vs_description = self.vs_description
				wafconfig.vs_port = self.vs_port
				wafconfig.vs_redirect = self.vs_redirect
				# wafconfig.vs_persistance = self.vs_persistance

				# add virtual server pool info to config object
				wafconfig.pool_name = self.vs_name
				wafconfig.pool_description = self.vs_description
				wafconfig.pool_loadBalancingMode = ''  # self.vs_pool_loadBalancingMode
				wafconfig.pool_monitor = ''  # self._pool_monitor
				# wafconfig.waf = waf  # TODO: is a dictionary from environments in settings. Need to class devices and environments

				self.configs.append(wafconfig)
				self.configs_waf.append(wafconfig)

	def _gen_gtm_configs(self):
		if 'WAF' in self.environment.keys():
			for waf in self.environment['WAF']:
				wafconfig = F5WAFConfig()

				# Add device information to config object
				wafconfig.d_type = 'waf'
				wafconfig.d_ip = waf['ip']
				wafconfig.d_name = waf['name'].lower()
				wafconfig.d_location = waf['location'].lower()
				wafconfig.d_prod = waf['prod'].lower()  # TODO: Create non prod name if this is a non-prod device
				wafconfig.d_gtmpool = waf['gtmpool']

				# add virtual server info to config object
				wafconfig.vs_name = self.vs_name
				wafconfig.vs_ip = self._get_ip(wafconfig)  # gets ip according to location and whether prod/non-prod
				wafconfig.vs_description = self.vs_description
				wafconfig.vs_port = self.vs_port
				wafconfig.vs_redirect = self.vs_redirect
				# wafconfig.vs_persistance = self.vs_persistance

				# add virtual server pool info to config object
				wafconfig.pool_name = self.vs_name
				wafconfig.pool_description = self.vs_description
				wafconfig.pool_loadBalancingMode = ''  # self.vs_pool_loadBalancingMode
				wafconfig.pool_monitor = ''  # self._pool_monitor
				# wafconfig.waf = waf  # TODO: is a dictionary from environments in settings. Need to class devices and environments

				self.configs.append(wafconfig)
				self.configs_waf.append(wafconfig)
			waflist = []
			for gtm in self.environment['GTM']:
				for waf in self.configs_waf:
					gtmconfig = F5GTMConfig()

					# Add device information to config object
					gtmconfig.d_type = 'gtm'
					gtmconfig.d_ip = gtm['ip']
					gtmconfig.d_name = gtm['name'].lower()
					gtmconfig.d_location = gtm['location'].lower()
					gtmconfig.d_prod = gtm['prod'].lower()  # TODO: Create non prod name if this is a non-prod device
					gtmconfig.d_gtmpool = gtm['gtmpool']

					gtmconfig.gtm = gtm
					gtmconfig.type = 'gtm'

					# NAME CHANGES / PORT CHANGES / POOL CHANGES

					gtmconfig.ltm_vs_port = waf.ltm_vs_port
					gtmconfig.gtmpool = waf.gtmpool
					gtmconfig.env = waf.env
					gtmconfig.gtm_pool_lb_method = 'global-availability'
					gtmconfig.gtm_wide_lb_method = 'global-availability'

					# NAME CHANGES / PORT CHANGES / POOL CHANGES


					gtmconfig.ltm_vs_name = waf.ltm_vs_name
					gtmconfig.waf_vip = waf.vs_ip
					gtm_members = []
					for w in self.configs_waf:
						if {'name': w.gtmpool + ":%s" % gtmconfig.ltm_vs_name} not in gtm_members:
							if str(waf.prod).lower() == str(w.prod).lower():
								print w.ltm_vs_name
								gtm_members.append({'name': str(w.gtmpool) + ":%s" % gtmconfig.ltm_vs_name})

					gtmconfig.destination = waf.waf_vip + ':' + gtmconfig.ltm_vs_port
					gtmconfig.gtm_members = gtm_members
					if str(waf.prod).lower() == 'prod':
						gtmconfig.wideip = gtmconfig.ltm_vs_name + ".gslb.barclaycardus.com"
					if str(waf.prod).lower() == 'non-prod':
						gtmconfig.wideip = gtmconfig.ltm_vs_name + ".gslb.barclaycardus.com"

					self.configs.append(gtmconfig)
					self.configs_gtm.append(gtmconfig)

		if 'GTM' in self.environment.keys() and 'WAF' not in self.environment.keys():
				ltm_list=[]
				for gtm in environment['GTM']:
					for ltm in ltm_data:

						gtm_members = []

						gtmdata = newdata()

						gtmdata.gtm = gtm
						gtmdata.type = 'gtm'

						#NAME CHANGES / PORT CHANGES / POOL CHANGES

						gtmdata.ltm_vs_port = ltm.ltm_vs_port
						gtmdata.gtmpool = ltm.gtmpool
						gtmdata.env = ltm.env
						gtmdata.environment = str(i[2]).strip()
						gtmdata.gtm_pool_lb_method = 'global-availability'
						gtmdata.gtm_wide_lb_method = 'global-availability'


						#NAME CHANGES / PORT CHANGES / POOL CHANGES


						gtmdata.ltm_vs_name = ltm.ltm_vs_name
						gtmdata.waf_vip = ltm.vip
						gtm_members=[]
						for l in ltm_data:
							if {'name': l.gtmpool + ":/Common/%s" % l.ltm_vs_name} not in gtm_members:
								if str(ltm.prod).lower() == str(l.prod).lower():
									gtm_members.append({'name': str(l.gtmpool) + ":/Common/%s" % l.ltm_vs_name})


						gtmdata.destination = ltm.vip +':'+ltm_vs_port
						gtmdata.gtm_members=gtm_members
						# if str(ltm.prod).lower() == 'prod':
						# 	gtmdata.wideip = gtmdata.ltm_vs_name + ".gslb.barclaycardus.com"
						# if str(ltm.prod).lower() == 'non-prod':
						# 	gtmdata.wideip = gtmdata.ltm_vs_name + ".gslb.barclaycardus.com"

						gtmdata.wideip=cdata.wideip

						if ltm.gtmpool not in ltm_list:
							ltm_list.append(ltm.gtmpool)
							modeled_data.append(gtmdata)



	def gen_configs(self):
		self._gen_ltm_configs()
		self._gen_waf_configs()

	def get_configs(self):
		return self.configs

	def enum_configs(self):
		return [config for config in enumerate(self.configs, start=1)]

	def get_config(self,idx):
		for config in self.enum_configs():
			if str(config[0])==str(idx):
				return config


	def configs_count(self):
		return self.configs.count()


class F5RequestCopy(F5Request):
	def __init__(self, LTMConfig):
		self.device_ip = ''
		self.name = ''
		self.description = ''
		self.port = ''
		self.protocol = ''
		self.redirection = ''
		self.persistance_profile = ''
		self.pool_name = ''
		self.pool_description = ''
		self.pool_loadBalancingMode = ''
		self.pool_monitor = ''


Option = namedtuple('Option', ['letter','label', 'callback'])


class RequestsMenu(RequestsHandler):
	mseperator =  "\n"* 82
	def __init__(self):
		super(RequestsMenu, self).__init__()

		self.requests = self.enumerate_requests()
		self.choice=''
		self.request=''
		self.config=''
		self.options=[]
		self.mainmenu = MainMenu("Make a selection:",[('Requests System({}) '.format(self.requests_count), self.requests_list_menu),('Import Requests', self.import_requests), ('Reporter', 'main_menu')])

	 	self.main_menu()
	#

		Option = namedtuple('Option', ['label', 'callback'])

	def main_menu(self):
		while True:

			self.mainmenu._options=[('Requests System({}) '.format(self.requests_count), self.requests_list_menu),('Import Requests', self.import_requests), ('Reporter', 'main_menu')]

			print self.mseperator
			print self.mainmenu.display()
			option = raw_input('\t>> ')

			try:
				option=int(option)
				self.mainmenu.callback(option)()
			except:
				continue

			# except Exception, e:
			# 	print("\t %s" % e)

	def requests_list_menu(self):
		print self.mseperator
		rlmenu = Menu("Make a selection:",[('c', '(C)onfigure All', 'config_all'), ('v', '(V)iew All Configs', 'all_configs_menu'),('o', '(O)ptions', 'config_options'),('b', '(B)ack', '')])

		while True:
			for request in self.enumerate_requests():
				print "\t{} {} implemented in the {} environment.".format(request[0], request[1].vs_name, request[1].env)

			print rlmenu.display()
			option = raw_input('\t>> ')

			try:
				option=int(option)
				request = self.get_request(option)
				self.request=request

			except Exception, e:
				option=str(option).lower()

			self.choice=option

			if type(self.choice)==int:
				self.request_menu()
			elif type(self.choice)==str:
				if self.choice.lower()=='b':
					return
				try:
					rlmenu.callback(option)()
				except:
					pass
			else:
				continue


	def import_requests(self,filename=''):
		super(RequestsMenu, self).import_requests()
		self.requests_list_menu()

	def request_menu(self):
		print self.mseperator
		request=self.request
		requestmenu = Menu("Make a selection:",[
			('c','(C)onfigure All', self.config_all),
			('v','(V)iew All Configs', self.generate_all_request_configs),
			('b','(B)ack', '')
		])

		while True:
			for c, config in enumerate(request.configs, start=1):
				print "\t{}) Configure {} {} {} {} with virtual server: {} and pool {} at destination:{}".format(
					c, config.d_name,config.d_location,config.d_prod,config.d_type, config.vs_name,config.pool_name,config.vs_ip + ':' + config.vs_port)


			print requestmenu.display()
			option = raw_input('\t>> ')

			if option.lower()=='b':
				return
			try:
				self.choice = int(option)
				self.configs_menu()
			except:
				pass

			try:
				str(option)
				requestmenu.callback(option)()
			except:
				pass
		return

	def configs_menu(self):
		print self.mseperator

		request = self.request
		configmenu = Menu("Make a selection:", [('p', 'Create (p)ool', self.config_all),
												('v', 'Create (v)irtual', self.generate_all_request_configs),
												('l', 'Delete virtua(l)', self.generate_all_request_configs),
												('o', 'Delete po(o)l', self.generate_all_request_configs),
												('r', '(R)eport virtual', self.generate_all_request_configs),
												('b', '(B)ack', self.request_menu)])
		while True:
			for c, config in enumerate(request.configs, start=1):
				if str(c) == str(self.choice):
					print "\n" *82
					print "\tDevice: {} \n\tLocation: {} \n\tProd: {} \n\tType: {}".format(config.d_name,config.d_location,config.d_prod,config.d_type)
					print "\n\tVirtual Server: {} \n\tPool: {} \n\tIP: {} \n\tPort: {}".format(config.vs_name,config.pool_name,config.vs_ip, config.vs_port)

					print configmenu.display()
					option = raw_input('\t>> ')


					try:
						str(option)
						configmenu.callback(option)()
					except:
						pass


	def all_configs_menu(self):
		print self.mseperator

		if len(self.configs_list) <=0 :
			self.generate_all_request_configs()

		print "\n" * 82

		#request = self.get_request(self.choice)
		allconfigsmenu = Menu("Make a selection:",
							  [
								  ('c', '(C)onfigure All', self.config_all),
								  ('b', '(B)ack',self.request_menu)])

		while True:
			for c, config in enumerate(self.configs_list, start=1):
				print "\t#{} Configure {} {} {} {} with virtual server: {} and pool {} at destination:{}".format(c,
																											   config.d_name,
																											   config.d_location,
																											   config.d_prod,
																											   config.d_type,
																											   config.vs_name,
																											   config.pool_name,
																											   config.vs_ip + ':' + config.vs_port)

			print allconfigsmenu.display()

			option = raw_input('\t>> ')




			try:
				str(option)
				allconfigsmenu.callback(option)()
			except:
				continue
		return







def dns_entry(name, zone='barcapint.com', type='A/PTR', crud='Create', desired_value='', comment='', current_value=''):
	existing_data = ''
	dns_record = [name, zone, type, crud, current_value, desired_value, comment]

	try:
		with open(r'dns_request.csv', 'r') as existing_data_file:
			reader = csv.reader(existing_data_file, delimiter=',', quotechar='|')
			existing_data = list(reader)
		# existing_data_file.close
		# print existing_data
	except IOError:
		pass
	with open(r'dns_request.csv', 'a') as f:
		if existing_data:
			if dns_record not in existing_data:
				print "DNS Entry added to DNS requests file."
				writer = csv.writer(f)
				writer.writerow(dns_record)
			else:
				print "DNS entry already exists in DNS requests file."
		else:
			writer = csv.writer(f)
			writer.writerow(dns_record)


x=RequestsMenu('envmix.xlsx')



#RequestsMenu()